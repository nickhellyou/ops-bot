import os
import telebot
from telebot import types
from openpyxl import load_workbook

TOKEN = os.getenv("TOKEN")
bot = telebot.TeleBot(TOKEN, parse_mode="HTML")


def main_keyboard():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.row("Kepala Sekolah", "Guru")
    kb.row("Siswa", "Kelas")
    kb.row("SPMB", "PPDB SMP")
    kb.row("Surat", "Dapodik")
    kb.row("Keuangan", "Tools")
    kb.row("Agenda", "BOS")
    return kb

def keyboard_surat():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)

    kb.row("Surat Tugas", "Surat Keterangan")
    kb.row("Surat Undangan", "Nomor Surat")
    kb.row("Template Surat")
    kb.row("Kembali")

    return kb


@bot.message_handler(commands=["start", "help"])
def start(message):
    bot.send_message(
        message.chat.id,
        "<b>OPS-BOT SDN 1 LANGSE</b>\n\nSilakan pilih menu.",
        reply_markup=main_keyboard()
    )


def balas(message, judul, isi):
    bot.send_message(
        message.chat.id,
        f"<b>{judul}</b>\n\n{isi}"
    )

def daftar_guru():
    wb = load_workbook("/storage/emulated/0/OPS-BOT/guru.xlsx")
    ws = wb.active

    hasil = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        nama = row[0]
        jabatan = row[1]

        if nama:
            hasil.append(
                f"{i}. {nama}\n"
                f"   Jabatan : {jabatan}"
            )

    if hasil:
        return "\n\n".join(hasil)
    else:
        return "Data guru masih kosong."

def daftar_siswa():
    from openpyxl import load_workbook

    try:
        wb = load_workbook("/storage/emulated/0/OPS-BOT/siswa.xlsx")
        ws = wb.active

        hasil = []

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            nama = row[0]
            kelas = row[1] if len(row) > 1 else ""

            if nama:
                hasil.append(
                    f"{i}. {nama}\n"
                    f"   Kelas : {kelas}"
                )

        if hasil:
            return "\n\n".join(hasil)
        else:
            return "Data siswa masih kosong."

    except Exception as e:
        return f"Error membaca data:\n{e}"

def menu_surat(message):
    bot.send_message(
        message.chat.id,
        "<b>MENU SURAT</b>\n\nSilakan pilih jenis surat.",
        reply_markup=keyboard_surat()
    )


@bot.message_handler(func=lambda m: True)
def menu(message):
    t = message.text

    if t == "Kepala Sekolah":
        balas(
            message,
            "KEPALA SEKOLAH",
            "Ahmad Junaedi Fatah, S.Pd.SD\n\nFitur profil sekolah akan dikembangkan."
        )

    elif t == "Guru":
        balas(
        message,
        "DATA GURU",
        daftar_guru()
    )

    elif t == "Siswa":
        balas(
        message,
        "DATA SISWA",
        daftar_siswa()
    )

    elif t == "Kelas":
        balas(
            message,
            "DATA KELAS",
            "Manajemen kelas akan dikembangkan."
        )

    elif t == "SPMB":
        balas(
            message,
            "SPMB",
            "Pendaftaran murid baru, murid pindahan, persyaratan, dan daftar ulang."
        )

    elif t == "PPDB SMP":
        balas(
            message,
            "PPDB SMP",
            "Jalur Domisili, Afirmasi, Prestasi, Persyaratan, dan Jadwal."
        )

    elif t == "Surat":
        menu_surat(message)

    elif t == "Surat Tugas":
        balas(
            message,
            "SURAT TUGAS",
            "Template Surat Tugas\n\nFitur generator surat akan segera tersedia."
        )

    elif t == "Surat Keterangan":
        balas(
            message,
            "SURAT KETERANGAN",
            "Template Surat Keterangan\n\nFitur generator surat akan segera tersedia."
        )

    elif t == "Surat Undangan":
        balas(
            message,
            "SURAT UNDANGAN",
            "Template Surat Undangan\n\nFitur generator surat akan segera tersedia."
        )

    elif t == "Nomor Surat":
        balas(
            message,
            "NOMOR SURAT",
            "Nomor Surat Terakhir:\n001/SDN1-LGS/VII/2026"
        )

    elif t == "Template Surat":
        balas(
            message,
            "TEMPLATE SURAT",
            "1. Surat Tugas\n"
            "2. Surat Keterangan\n"
            "3. Surat Undangan\n"
            "4. Surat Izin\n"
            "5. Surat Pemberitahuan"
        )

    elif t == "Kembali":
        bot.send_message(
            message.chat.id,
            "<b>OPS-BOT SDN 1 LANGSE</b>\n\nSilakan pilih menu.",
            reply_markup=main_keyboard()
        )

    elif t == "Dapodik":
        balas(
            message,
            "DAPODIK",
            "Informasi Dapodik, sinkronisasi, dan link penting."
        )

    elif t == "Keuangan":
        balas(
            message,
            "KEUANGAN",
            "Dana BOS, Dana Lain-lain, Pemasukan, Pengeluaran, dan Laporan."
        )

    elif t == "Agenda":
        balas(
            message,
            "AGENDA SEKOLAH",
            "Jadwal lomba, ANBK, TKA, rapat, dan agenda sekolah."
        )

    elif t == "Tools":
        balas(
            message,
            "TOOLS OPERATOR",
            "BMD\nGenerator Surat\nNomor Surat\nBackup Database\nExport PDF"
        )

    elif t == "BOS":
        balas(
            message,
            "DANA BOS",
            "RKAS, Realisasi, Laporan, dan Berkas BOS."
        )

    elif t == "Kembali":
        bot.send_message(
            message.chat.id,
            "<b>OPS-BOT SDN 1 LANGSE</b>\n\nSilakan pilih menu.",
            reply_markup=main_keyboard()
        )

    else:
        balas(
            message,
            "OPS-BOT",
            "Silakan pilih menu yang tersedia."
        )


print("OPS-BOT SDN 1 LANGSE berjalan...")
bot.infinity_polling()
